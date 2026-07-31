"""데이터 변환 — CSV↔XLSX, CSV↔JSON (REQ-F-006). 한글 인코딩 자동 감지 (REQ-F-009 Should)."""
import csv
import datetime
import io
import json
from pathlib import Path

from .base import ConversionError

_ENCODINGS = ("utf-8-sig", "cp949", "utf-8")


def _read_text(path: Path) -> str:
    for enc in _ENCODINGS:
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
        except OSError:
            raise ConversionError("err.corrupted")
    raise ConversionError("err.encoding")


def _read_csv_rows(path: Path) -> list[list[str]]:
    text = _read_text(path)
    # 구분자(콤마/세미콜론/탭)만 자동 감지한다. 셀 안의 줄바꿈·이스케이프된
    # 큰따옴표(")까지 함께 있는 실제 데이터에서 csv.Sniffer가 quotechar나
    # doublequote까지 함께 추측하면 오탐이 잦아 값이 중간에서 잘리는 문제가
    # 있었다 — 따옴표 규칙은 표준(RFC4180, csv.excel 기본값)으로 고정한다.
    # text.splitlines()로 미리 줄 단위로 쪼개면 셀 안의 줄바꿈이 있는 필드가
    # 별도 행으로 잘못 쪼개지므로, csv.reader에 원문 텍스트를 그대로 넘긴다.
    try:
        delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t").delimiter
    except csv.Error:
        delimiter = ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def csv_to_xlsx(src: Path, tmpdir: Path) -> Path:
    from openpyxl import Workbook
    rows = _read_csv_rows(src)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for row in rows:
        ws.append(row)
    out = tmpdir / (src.stem + ".xlsx")
    wb.save(out)
    return out


def _format_cell_value(v):
    """엑셀에서 보이는 모습에 맞춰 값을 문자열로 정규화한다.

    openpyxl은 날짜 셀을 datetime 객체로, 정수처럼 보이는 셀도 내부적으로
    float(예: 3.0)로 돌려줄 수 있다 — csv.writer가 그대로 str()하면
    "2026-07-31 00:00:00"·"3.0"처럼 엑셀에서 보던 모습과 달라져 "이상하게
    추출됐다"고 느끼기 쉽다. 날짜는 ISO 형식(시간이 자정이 아닐 때만 시간
    포함), 정수값 float은 소수점 없이 표시한다.
    """
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        if v.time() == datetime.time(0, 0):
            return v.date().isoformat()
        return v.isoformat(sep=" ")
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, datetime.time):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return v


def xlsx_sheet_count(path: Path) -> int:
    """워크북의 시트 개수. 열 수 없으면(손상 등) 1로 본다 — 실제 변환 시도에서
    본연의 오류(err.corrupted 등)로 다시 드러나므로 여기서는 UI 고지 판단용."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, read_only=True)
        n = len(wb.sheetnames)
        wb.close()
        return n
    except Exception:
        return 1


def xlsx_to_csv(src: Path, tmpdir: Path) -> Path:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(src, read_only=True, data_only=True)
    except Exception as e:  # zip 손상, 암호화 등
        key = "err.password" if "encrypt" in str(e).lower() else "err.corrupted"
        raise ConversionError(key, str(e))
    # 시트가 여러 개면 첫 번째(활성) 시트만 변환한다 — 나머지 시트를 조용히
    # 버리지 않도록 변환 전 UI에 고지한다(note.xlsx_multisheet, main_window.py).
    # 여러 시트를 각각 파일로 출력하는 방안은 "입력 1개 → 출력 1개" 전제인
    # 현재 데이터 모델(FileItem.output)을 바꿔야 해서 별도 과제로 보류.
    ws = wb.active
    out = tmpdir / (src.stem + ".csv")
    # utf-8-sig: 엑셀에서 한글 깨짐 없이 열리도록 BOM 포함
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([_format_cell_value(v) for v in row])
    wb.close()
    return out


def csv_to_json(src: Path, tmpdir: Path) -> Path:
    rows = _read_csv_rows(src)
    if not rows:
        records: list = []
    else:
        header, body = rows[0], rows[1:]
        records = [dict(zip(header, row)) for row in body]
    out = tmpdir / (src.stem + ".json")
    out.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def json_to_csv(src: Path, tmpdir: Path) -> Path:
    try:
        payload = json.loads(_read_text(src))
    except json.JSONDecodeError as e:
        raise ConversionError("err.corrupted", str(e))
    if not isinstance(payload, list):
        raise ConversionError("err.jsonshape")

    out = tmpdir / (src.stem + ".csv")
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if all(isinstance(r, dict) for r in payload):
            keys: list[str] = []
            for r in payload:
                for k in r:
                    if k not in keys:
                        keys.append(k)
            writer.writerow(keys)
            for r in payload:
                writer.writerow([r.get(k, "") for k in keys])
        elif all(isinstance(r, list) for r in payload):
            writer.writerows(payload)
        else:
            raise ConversionError("err.jsonshape")
    return out
