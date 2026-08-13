"""FileRow 단순화 고지 문구 테스트 — DEC-010(PDF/HWP→DOCX)·DEC-017(DOCX→HWP)·
XLSX→CSV 다중 시트.

리뷰 지적(코드 리뷰): DOCX→HWP는 표 구조가 손실되는 신규 경로인데 변환 전
UI 고지가 기존 DEC-010 범위로 확장되지 않았음 — 보완 후 회귀 방지용.
"""
import os
import shutil
import tempfile
import unittest
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication

from app import i18n
from app.models import FileItem
from app.tokens import LIGHT as TOKENS
from app.ui.main_window import FileRow

_app = QApplication.instance() or QApplication([])


class TestFormatNote(unittest.TestCase):
    def setUp(self):
        # tr()은 시스템 로케일을 따른다(DEC-009) — CI 러너 로케일에 좌우되지
        # 않도록 언어를 한국어로 고정한다. QSettings에 실제 저장되므로
        # (로컬 실행 시 앱의 실제 언어 설정에 영향) 원래 값을 복원한다.
        self._orig_lang_pref = i18n.saved_pref()
        i18n.set_lang("ko")
        self.tmp = Path(tempfile.mkdtemp())

    def tearDown(self):
        i18n.set_lang(self._orig_lang_pref or None)
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _note_for(self, source_fmt: str, target_fmt: str, source_path: Path | None = None):
        item = FileItem(id=1, source=source_path or Path(f"test.{source_fmt}"),
                         source_fmt=source_fmt, target_fmt=target_fmt)
        row = FileRow(item, TOKENS, lambda i: None, lambda: None)
        row._update_note()
        return (not row.reason.isHidden()), row.reason.text()

    def test_docx_to_hwp_shows_layout_simplified_note(self):
        """DEC-028: 표는 이제 실제 HWP 표로 만들어지지만(DEC-017 정정),
        셀 병합·정밀한 레이아웃까지는 아니라 여전히 단순화 고지가 뜬다."""
        visible, text = self._note_for("docx", "hwp")
        self.assertTrue(visible)

    def test_pdf_to_docx_shows_absolute_position_note(self):
        """DEC-037: PDF→DOCX가 줄 단위 절대 위치 재구성으로 바뀌면서
        전용 고지(note.pdf_to_docx)로 갈아탔다 — 편집 시 줄이 자연스럽게
        안 이어질 수 있다는 트레이드오프 안내가 핵심. 이미지·표 테두리
        반영 개선 이후로는 그 둘도 "이제 반영된다"는 쪽으로 문구가
        바뀌었다(예전엔 반대로 "옮겨지지 않는다"는 안내였음)."""
        visible, text = self._note_for("pdf", "docx")
        self.assertTrue(visible)
        self.assertIn("고정된 위치", text)
        self.assertIn("이미지", text)
        self.assertIn("표 테두리", text)

    def test_hwp_to_docx_shows_layout_simplified_note(self):
        visible, text = self._note_for("hwp", "docx")
        self.assertTrue(visible)

    def test_pdf_to_hwp_shows_layout_simplified_note(self):
        visible, text = self._note_for("pdf", "hwp")
        self.assertTrue(visible)

    def test_hwpx_to_docx_shows_layout_simplified_note(self):
        """HwpxToJson.java가 이제 머리말·꼬리말·글상자까지 재귀로 훑는다
        (hwplib 쪽 DEC-032와 대칭, 표 셀 안 서식 보존에 이은 개선) — 전용
        note.hwpx_simplified가 더는 필요 없어져 HWP와 같은 일반
        note.simplified로 통합됐다."""
        visible, text = self._note_for("hwpx", "docx")
        self.assertTrue(visible)

    def test_hwpx_to_pdf_shows_layout_simplified_note(self):
        visible, text = self._note_for("hwpx", "pdf")
        self.assertTrue(visible)

    def test_docx_to_pdf_shows_no_note(self):
        visible, _ = self._note_for("docx", "pdf")
        self.assertFalse(visible)

    def test_hwp_to_pdf_shows_no_note(self):
        visible, _ = self._note_for("hwp", "pdf")
        self.assertFalse(visible)

    def test_xlsx_to_csv_multisheet_shows_note(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(["a"])
        wb.create_sheet("두번째")
        src = self.tmp / "multi.xlsx"
        wb.save(src)
        visible, text = self._note_for("xlsx", "csv", source_path=src)
        self.assertTrue(visible)
        self.assertIn("시트", text)

    def test_xlsx_to_csv_single_sheet_shows_no_note(self):
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(["a"])
        src = self.tmp / "single.xlsx"
        wb.save(src)
        visible, _ = self._note_for("xlsx", "csv", source_path=src)
        self.assertFalse(visible)

    def test_animated_gif_to_png_shows_first_frame_note(self):
        from PIL import Image
        frames = [Image.new("RGB", (4, 4), (i * 50, 0, 0)) for i in range(3)]
        src = self.tmp / "anim.gif"
        frames[0].save(src, save_all=True, append_images=frames[1:], duration=50, loop=0)
        visible, text = self._note_for("gif", "png", source_path=src)
        self.assertTrue(visible)
        self.assertIn("첫 번째", text)

    def test_static_png_to_jpg_shows_no_note(self):
        from PIL import Image
        src = self.tmp / "static.png"
        Image.new("RGB", (4, 4)).save(src)
        visible, _ = self._note_for("png", "jpg", source_path=src)
        self.assertFalse(visible)

    def test_pdf_to_images_shows_folder_note(self):
        """DEC-026 — PNG. DEC-043으로 JPG도 같은 고지를 받아야 한다."""
        visible, text = self._note_for("pdf", "png")
        self.assertTrue(visible)
        self.assertIn("폴더", text)

    def test_pdf_to_jpg_shows_folder_note(self):
        visible, text = self._note_for("pdf", "jpg")
        self.assertTrue(visible)
        self.assertIn("폴더", text)

    def test_pdf_to_pptx_shows_layout_note(self):
        """DEC-030/DEC-036: 텍스트·이미지·표 테두리 모두 원래 위치로 재구성되지만
        복잡한 곡선 도형은 다각형으로 근사된다는 고지 — note.simplified(단순화)와는
        다른 문구."""
        visible, text = self._note_for("pdf", "pptx")
        self.assertTrue(visible)
        self.assertIn("표 테두리", text)
        self.assertIn("근사", text)

    def test_obj_to_stl_shows_no_color_note(self):
        """STL 포맷 자체가 색상/재질을 못 담아 다른 3D 포맷 → STL 변환에서만 뜬다."""
        visible, text = self._note_for("obj", "stl")
        self.assertTrue(visible)
        self.assertIn("색상", text)

    def test_obj_to_glb_shows_no_note(self):
        """STL이 아닌 3D 포맷 간 변환은 색상·재질이 보존되므로 고지 없음."""
        visible, text = self._note_for("obj", "glb")
        self.assertFalse(visible)


if __name__ == "__main__":
    unittest.main()
